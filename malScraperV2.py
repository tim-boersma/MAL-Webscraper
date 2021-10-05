import requests
import re
import csv
import columnNames
import datetime
import os.path
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font

#KNOWN BUGS:
#Updates to a file will not write index titles if there wasn't a show in the index when the file was created
#Incorrectly places new title if it is the most popular in its index
def updateColumnA(filename, URL):

    if os.path.isfile(filename) == True:
        updateOpenColumn(filename, URL)
        return

    #webscraper setup
    page = requests.get(URL)
    soup = BeautifulSoup(page.content, 'html.parser')
    results = soup.find(id="content")   
    titles_elem = results.find_all('a', class_="link-title")

    #excel sheet setup
    workbook = Workbook()
    scoreSheet = workbook.active
    workbook.create_sheet('Members')
    memberSheet = workbook['Members']
    scoreSheet.title = 'Scores'
    scoreSheet["A1"] = memberSheet["A1"] = "Title"

    #variable declaration
    createIndex(URL)
    titleCount = 2
    longestTitle = 0;
    global animeIndex
    global indexTitles

    for title in titles_elem:
        #Bolds text if it's an index
        for i in range(len(animeIndex)):
            if title.text.strip() == animeIndex[i]:
                scoreSheet["A" + str(titleCount)] = memberSheet["A" + str(titleCount)] = indexTitles[i]
                scoreSheet["A" + str(titleCount)].font = Font(bold = True)
                memberSheet["A" + str(titleCount)].font = Font(bold = True)
                titleCount += 1
        #Writes title to score and member sheet
        scoreSheet["A" + str(titleCount)] = memberSheet["A" + str(titleCount)] = title.text.strip()
        
        #Saves text length if longer than longest title
        if len(str(title.text.strip())) > longestTitle:
            longestTitle = len(str(title.text.strip()))
        titleCount += 1

    #prevents longest title from being too long
    if longestTitle > 70:
        longestTitle = 70

    #Sets title column width
    scoreSheet.column_dimensions["A"].width = longestTitle
    memberSheet.column_dimensions["A"].width = longestTitle

    #saves the file
    workbook.save(filename=filename)
    
    #Updates files to input scores and member counts 
    updateOpenColumn(filename, URL)

def updateOpenColumn(filename, URL):
    if os.path.isfile(filename) == False:
        print("File does not exist. Creating new file...")
        updateColumnA(filename, URL)
        return
    
    #index setup
    createIndex(URL)
    global columnList
    columnList = columnNames.getColumnNames()
    global animeIndex
    date = datetime.datetime.now().strftime("%m/%d/%Y")

    #webscraper setup
    page = requests.get(URL)
    soup = BeautifulSoup(page.content, 'html.parser')
    results = soup.find(id="content")  

    #excel sheet setup
    workbook = load_workbook(filename)
    scoreSheet = workbook['Scores']
    memberSheet = workbook['Members']

    #find open column
    openScoreCol = findOpenColumn(scoreSheet)
    openMemberCol = findOpenColumn(memberSheet)

    #set headers
    scoreSheet[str(columnList[openScoreCol]) + "1"] = str(date)
    memberSheet[str(columnList[openMemberCol]) + "1"] = str(date)
    scoreSheet.column_dimensions[str(columnList[openScoreCol])].width = 10
    memberSheet.column_dimensions[str(columnList[openMemberCol])].width = 10

    #create titleList 
    titleList = []
    titles_elem = results.find_all('a', class_="link-title")
    for title in titles_elem:
        titleList.append(title.text.strip())
        
    index = 0
    scores_elem = results.find_all('span', class_=re.compile("score[a-z \-A-Z\d]*"))
    for score in scores_elem:
        #Find where to put title
        titleIndex = findTitle(titleList[index], scoreSheet)
        #make space if it is not in the spreadsheet and fills in previous score sections with '-'
        if titleIndex == -1:
            titleIndex = insertNewTitle(scoreSheet, titleList[index - 1], titleList[index])
            
        #write score as a string if it's N/A
        if score.text.strip() == "N/A":
            scoreSheet[str(columnList[openScoreCol]) + str(titleIndex)] = score.text.strip()
        #write score as a float if its a number
        else:
            scoreSheet[str(columnList[openScoreCol]) + str(titleIndex)] = float(score.text.strip())
        #increment index
        index += 1

    index = 0
    members_elem = results.find_all('span', class_="member fl-r")
    for member in members_elem:
         
        #Find where to put title
        titleIndex = findTitle(titleList[index], memberSheet)
        #make space if it is not in the spreadsheet and fills in previous member sections with '-'
        if titleIndex == -1:
            titleIndex = insertNewTitle(memberSheet, titleList[index - 1], titleList[index])

        #write member to excel sheet
        memberSheet[str(columnList[openMemberCol]) + str(titleIndex)] = int(member.text.strip().replace(',', ''))
        #increment index
        index += 1

    #sort sheets
    sortSheet(3, scoreSheet)
    sortSheet(3, memberSheet)

    #removes titles if they were not updated
    rowCleanup(scoreSheet)
    rowCleanup(memberSheet)

    #saves workbook
    workbook.save(filename=filename)    

def rowCleanup(sheet):
    column = findOpenColumn(sheet) - 1
    bottomRow = index = 3

    #get bottom row
    while not sheet['A' + str(bottomRow + 1)].value is None:
        bottomRow += 1
    
    #decrement down the rows
    while sheet[columnList[column] + str(index)].row <= bottomRow:
        #if the last data column is empty in a row and the column of A the same row is not bold then you know the entry was deleted and it's not a title index
        if sheet[columnList[column] + str(index)].value is None and not sheet["A" + str(index)].font.bold:
            currentRow = sheet[columnList[column] + str(index)].row
            #move all rows beneath the deleted entry up 1
            while currentRow <= bottomRow: 
                #copies the value of current row from the row beneath it
                for i in range(column + 1):
                    sheet[columnList[i] + str(currentRow)].value = sheet[columnList[i] + str(currentRow + 1)].value
                
                #change bolded font if needed
                if sheet["A" + str(currentRow + 1)].font.bold:
                    sheet["A" + str(currentRow + 1)].font = Font(bold = False)
                    sheet["A" + str(currentRow)].font = Font(bold = True)
                #move down the sheet
                currentRow += 1
            #once the last row is reached it is deleted because it has already been copied
            for i in range(column):
                sheet[columnList[i] + str(currentRow)].value = None
            bottomRow -= 1
        index += 1    

def sortSheet(min, sheet):
    global columnList
    max = min
    updated = False
    while str(sheet["A" + str(max)].value) != "None" and not sheet["A" + str(max)].font.bold:
        max += 1
    max -= 1
    index1 = min
    currentCol = findOpenColumn(sheet) - 1
    rowValues = [None] * (currentCol + 1)

    while index1 < max:
        index2 = min

        while index2 < max:
            if sheet[columnList[currentCol] + str(index2)].value == "N/A" or sheet[columnList[currentCol] + str(index2)].value is None:
                value1 = -1
            else: 
                value1 = float(sheet[columnList[currentCol] + str(index2)].value)
            if sheet[columnList[currentCol] + str(index2 + 1)].value == "N/A" or sheet[columnList[currentCol] + str(index2 + 1)].value is None:
                value2 = -1
            else: 
                value2 = float(sheet[columnList[currentCol] + str(index2 + 1)].value)

            if value1 < value2:
                updated = True
                for i in range(currentCol + 1):
                    rowValues[i] = sheet[columnList[i] + str(index2)].value

                for i in range(currentCol + 1):
                    sheet[columnList[i] + str(index2)].value = sheet[columnList[i] + str(index2 + 1)].value
                for i in range(currentCol + 1):
                    sheet[columnList[i] + str(index2 + 1)].value = rowValues[i]

            index2 += 1

        if not updated:
            break
        index1 += 1
    if str(sheet['A' + str(max + 2)].value) != "None":
        sortSheet(max + 2, sheet)

# returns the row if the title exists in the file or -1 if it does not
def findTitle(title, sheet):
    global indexTitles
    titleIndex = 2
    cell = sheet['A' + str(titleIndex)]
    while str(cell.value) != "None":
        if str(cell.value) == title:
            return titleIndex
        titleIndex += 1
        cell = sheet['A' + str(titleIndex)]
    return -1

def findOpenColumn(sheet):
    global columnList
    openCol = 1
    while str(sheet[str(columnList[openCol]) + "1"] .value) != "None":
        openCol += 1
    return openCol

def insertNewTitle(sheet, previousTitle, title):
       
    openCol = findOpenColumn(sheet) 

    titleIndex = findTitle(previousTitle, sheet) + 1

    while str(sheet["A" + str(titleIndex)].value) != "None" and not sheet["A" + str(titleIndex)].font.bold:
        titleIndex += 1

    openCellSpace(titleIndex, sheet)
    sheet['A' + str(titleIndex)] = title

    #set previous scores to 'N/A'
    columnIndex = 1
    while sheet[columnList[columnIndex] + str(titleIndex)].column - 1 != openCol:
        sheet[columnList[columnIndex] + str(titleIndex)].value = 'N/A'
        columnIndex += 1

    return titleIndex

#move bold cells
def openCellSpace(row, sheet):
    #variable declarationw
    global columnList
    colNum = findOpenColumn(sheet)
    cellValues = [None] * colNum

    #go to the bottom row
    cell = sheet["A1"]
    currentRow = 1
    while str(cell.value) != "None":
        currentRow += 1
        cell = sheet['A' + str(currentRow)]
    
    #shift all rows down 1
    while currentRow > row:
       
        #copies the value of current row from the row beneath it
        for i in range(colNum + 1):
            sheet[columnList[i] + str(currentRow)].value = sheet[columnList[i] + str(currentRow - 1)].value

        #change bolded font if needed
        if sheet["A" + str(currentRow - 1)].font.bold:
            sheet["A" + str(currentRow - 1)].font = Font(bold = False)
            sheet["A" + str(currentRow)].font = Font(bold = True)
        #move up the sheet
        currentRow -= 1
    #once the specified row is reached it is deleted because it has already been copied
    for i in range(colNum):
        sheet[columnList[i] + str(currentRow)].value = None


def createIndex(URL):
    page = requests.get(URL)
    soup = BeautifulSoup(page.content, 'html.parser')
    results = soup.find(id="content")
    global indexTitles
    indexTitles = ['TV (New)', 'TV (Continuing)', 'ONA', 'OVA', 'Movie', 'Special']
    global animeIndex
    global currentSeason
    animeIndex = [] 

    addIndex = False
    indexCount = seasonIndex = 3
    for text in soup.stripped_strings:
        if len(animeIndex) == 6:
            break
        if addIndex == True and text != "Watch Promotional Video" and text != "Watch Video":
            animeIndex.append(str(text))
            addIndex = False
        elif seasonIndex == 2:
            currentSeason = text[(len(text) - 4):] + text[:-5]
        elif text == '...':
            seasonIndex = 0
        elif str(text) == indexTitles[len(animeIndex)]:
            addIndex = True
        elif str(text) == 'ONA' and len(animeIndex) == 1:
            animeIndex.append('-')
            addIndex = True
        indexCount += 1
        seasonIndex += 1
    animeIndex.append("None")

def intTryParse(value):
    try:
        int(value)
        return True
    except ValueError:
        return False    


updateOpenColumn("C:\\Users\\Tim\\OneDrive - Dakota State University\\Webscraper\\2021spring.xlsx", "https://myanimelist.net/anime/season/2021/spring")
updateOpenColumn("C:\\Users\\Tim\\OneDrive - Dakota State University\\Webscraper\\2021summer.xlsx", "https://myanimelist.net/anime/season/2021/summer")
exit()

year = input('What year would you like to look at? >')
if year == "":
    URL = "https://myanimelist.net/anime/season"
else:
    
    if intTryParse(year) == False or int(year) > 2021 or int(year) < 1960:
        print('Invalid Year')
        exit()

    elif int(year) >= 1950 or int(year) <= 2021:
        season = input("What season would you like? (winter/spring/summer/fall) >")
        if season == "winter" or season == '1':
            URL = "https://myanimelist.net/anime/season/" + str(year) + "/winter"
        elif season == "spring" or season == '2':
            URL = "https://myanimelist.net/anime/season/" + str(year) + "/spring"
        elif season == "summer" or season == '3':
            URL =  "https://myanimelist.net/anime/season/" + str(year) + "/summer"
        elif season == "fall" or season == '4':
            URL = "https://myanimelist.net/anime/season/" + str(year) + "/fall"
        else:
            print('Invalid Season')
            exit()
    else:
        print("Invalid Year")
        exit()

if year == "":
    file = "C:\\Users\\Tim\\OneDrive - Dakota State University\\Webscraper\\" + str(currentSeason.lower()) + ".xlsx"
else:
    season = URL[42:]
    file = "C:\\Users\\Tim\\OneDrive - Dakota State University\\Webscraper\\" + str(year) + str(season) + ".xlsx"

updateColumnA(file, URL)